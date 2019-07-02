from openpyxl import Workbook
from openpyxl.drawing.image import Image as Im
import datetime
from PIL import Image
import os
import psutil
from functions import sql_connection


height = 100
width = 100


if not os.path.exists(os.getcwd() + '/report_xls'):
    os.mkdir(os.getcwd() + '/report_xls')


def has_handle(fpath):
    for proc in psutil.process_iter():
        try:
            for item in proc.open_files():
                if fpath == item.path:
                    return True
        except:
            pass

    return False


def is_in_use(fpath):
    while has_handle(fpath):
        print("in use")


def generate_xls(id_type, i_d, from_dt, to_dt):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for row in ws.iter_rows():
        for cell in row:
            cell.style.alignment.wrap_text = True

    result = sql_query(id_type, i_d, from_dt, to_dt)
    if len(result) == 0:
        return 0
    no_entries = len(result)

    ws['A1'] = 'Serial No'
    ws['B1'] = 'DateTime'
    ws['C1'] = 'Duplicate'
    ws['D1'] = 'Times Visited'
    ws['E1'] = 'Photo'

    for row in range(2, no_entries + 2):
        colname = 'A' + str(row)
        ws[colname] = str(row-1)

        colname = 'B' + str(row)
        ws[colname] = str(result[row-2]['timecapture'])
        ws.column_dimensions['B'].width = 20

        colname = 'C' + str(row)
        ws[colname] = result[row-2]['isduplicate']

        colname = 'D' + str(row)
        ws[colname] = result[row-2]['timesvisited']
        length = len(result[row-2]['path'])
        if length != 0:
            img_path_rel = result[row-2]['path']
            img_path_abs = '/home/sparsh/Desktop/Trial' + img_path_rel
            temp_path = os.getcwd()+'/temp' + str(row) + '.jpg'
            img = Image.open(img_path_abs)
            img = img.resize((width, height), Image.NEAREST)
            img.save(temp_path)
            is_in_use(temp_path)
            img_xls = Im(temp_path)
            print(type(img_xls))
            colname = 'E' + str(row)
            ws.column_dimensions['E'].width = 15
            ws.row_dimensions[row].height = 80
            is_in_use(temp_path)
            ws.add_image(img_xls, colname)
            is_in_use(temp_path)

    file_name = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    wb.save('report_xls/' + file_name + '.xlsx')
    dir_list = os.listdir(os.getcwd())
    for i in range(0, len(dir_list)):
        if '.jpg' in dir_list[i]:
            os.remove(os.getcwd() + '/' + dir_list[i])

    return file_name + '.xlsx'


def sql_query(id_type, i_d, from_dt, to_dt):
    connection = sql_connection()[1]
    with connection.cursor() as cursor:
        query = "SELECT * from tx_face_id where {} = '{}'" \
                 " and timecapture between '{}' and '{}'".format(id_type, i_d, from_dt, to_dt)
        cursor.execute(query)
        result = cursor.fetchall()
        print('hey')
        return result
